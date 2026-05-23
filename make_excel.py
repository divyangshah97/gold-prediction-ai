"""
Reads signals/signals.csv and produces a colour-coded signals.xlsx.
Run: python make_excel.py
Requires: pip install openpyxl
"""

import csv
import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Install openpyxl first:  pip install openpyxl")

SIGNALS_CSV = os.path.join(os.path.dirname(__file__), "signals", "signals.csv")
OUTPUT_XLS  = os.path.join(os.path.dirname(__file__), "signals", "signals.xlsx")

SIGNAL_COLORS = {
    "strong buy":  ("1A5C1A", "FFFFFF"),   # dark green, white text
    "buy":         ("4CAF50", "FFFFFF"),   # medium green, white text
    "wait":        ("FFC107", "212121"),   # amber, dark text
    "sell":        ("EF5350", "FFFFFF"),   # medium red, white text
    "strong sell": ("7B1010", "FFFFFF"),   # dark red, white text
}

HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
THIN         = Side(style="thin", color="CBD5E1")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_csv():
    rows = []
    with open(SIGNALS_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gold Signals"

    headers = ["Date", "Signal", "Score", "Reasoning"]
    col_widths = [14, 16, 8, 80]

    # Header row
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        signal_key = r.get("Signal", "").strip().lower()
        bg, fg = SIGNAL_COLORS.get(signal_key, ("FFFFFF", "212121"))
        signal_fill = PatternFill("solid", fgColor=bg)
        signal_font = Font(name="Calibri", size=10, bold=True, color=fg)

        values = [
            r.get("Date", ""),
            r.get("Signal", ""),
            r.get("Score", ""),
            r.get("Reasoning", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == 4),
                horizontal="center" if col_idx in (1, 3) else "left",
            )
            if col_idx == 2:
                cell.fill = signal_fill
                cell.font = signal_font
            else:
                cell.font = BODY_FONT

        ws.row_dimensions[row_idx].height = 40

    # Auto-filter on header
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

    wb.save(OUTPUT_XLS)
    print(f"Saved {len(rows)} signal(s) → {OUTPUT_XLS}")


if __name__ == "__main__":
    rows = load_csv()
    write_excel(rows)
